# backend/data_ingestion/spreadsheet_parser.py
"""
Parses coach-authored training plans from Excel/CSV formats.
Supports Layout A (weeks as rows), Layout B (weeks as columns), and Layout C (flat dates).
"""

import csv
import logging
import re
from datetime import date, timedelta
from pathlib import Path
from typing import Dict, List, Optional, Tuple

import openpyxl

from backend.storage.postgres_client import db

logger = logging.getLogger(__name__)

SPREADSHEET_IMPORT_DIR = Path("/data/imports/spreadsheets")
SPREADSHEET_IMPORT_DIR.mkdir(parents=True, exist_ok=True)

# ---------------------------------------------------------------------------
# Core Ingest Logic
# ---------------------------------------------------------------------------
def ingest_spreadsheet_plan(
    file_path: Path,
    athlete_id: str,
    plan_start_date: date,
    column_map: Optional[Dict] = None
) -> List[Dict]:
    """
    Main entry point. Detects layout, parses sessions, returns unified objects.
    plan_start_date: Monday of Week 1 (required for Layout A/B).
    """
    if file_path.suffix in (".xlsx", ".xlsm"):
        rows, headers = load_xlsx(file_path)
    elif file_path.suffix == ".csv":
        rows, headers = load_csv(file_path)
    else:
        raise ValueError(f"Unsupported spreadsheet format: {file_path.suffix}")

    layout = detect_layout(headers, rows)
    logger.info("Detected spreadsheet layout %s for %s", layout, file_path.name)

    if layout == "A":
        sessions = parse_layout_a(rows, headers, plan_start_date)
    elif layout == "C":
        sessions = parse_layout_c(rows, headers, column_map)
    elif layout == "B":
        sessions = parse_layout_b(rows, headers, plan_start_date)
    else:
        logger.warning("Falling back to LLM parsing for %s", file_path.name)
        # Note: Implement LLM parsing fallback here if deterministic fails
        sessions = []

    normalised = [_normalise_spreadsheet_session(s, athlete_id, file_path.name) for s in sessions]
    
    # Store to Database
    # Uses the upsert_planned_session method from the actual postgres client implementation 
    # but we can rely on the standard execute here too.
    for s in normalised:
        _store_planned_session(s)
        
    return normalised


def load_xlsx(path: Path) -> Tuple[List[List], List[str]]:
    """Load all sheets or the first non-empty sheet. Returns (rows, headers)."""
    wb = openpyxl.load_workbook(path, data_only=True)
    all_sessions = []
    headers = []
    
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        sheet_rows = [[cell.value for cell in row] for row in ws.iter_rows()]
        # Skip completely empty sheets
        if not any(any(c for c in r) for r in sheet_rows):
            continue
        # First non-empty row is the header
        for i, row in enumerate(sheet_rows):
            if any(c for c in row):
                if not headers:
                    headers = [str(c) if c else "" for c in row]
                    all_sessions.extend(sheet_rows[i+1:])
                else:
                    all_sessions.extend(sheet_rows[i:])
                break
                
    return all_sessions, headers


def load_csv(path: Path) -> Tuple[List[List], List[str]]:
    """Load CSV into (rows, headers)."""
    with open(path, encoding='utf-8') as f:
        reader = csv.reader(f)
        headers = next(reader, [])
        rows = list(reader)
    return rows, headers


# ---------------------------------------------------------------------------
# Layout Detectors and Parsers
# ---------------------------------------------------------------------------
def detect_layout(headers: List[str], rows: List[List]) -> str:
    """Sniff layout from header row shape."""
    h = [str(x).lower().strip() for x in headers if x]
    DAY_NAMES = {"mon","tue","wed","thu","fri","sat","sun",
                 "monday","tuesday","wednesday","thursday","friday","saturday","sunday"}
    
    # Needs at least Mon/Wed/Fri etc to be considered Layout A
    if len(DAY_NAMES & set(h)) >= 4:
        return "A"
    
    if "date" in h and any(x in h for x in ("sport","activity","type","discipline")):
        return "C"
        
    week_in_header = any("wk" in x or "week" in x for x in h)
    
    first_col = [str(r[0]).lower() for r in rows[:8] if r and r[0]]
    sport_in_first_col = any(
        any(s in v for s in ("run","swim","bike","ride","strength"))
        for v in first_col
    )
    
    if week_in_header and sport_in_first_col:
        return "B"
        
    return "unknown"


DAY_OFFSETS = {
    "mon":0,"monday":0,"tue":1,"tuesday":1,"wed":2,"wednesday":2,
    "thu":3,"thursday":3,"fri":4,"friday":4,"sat":5,"saturday":5,"sun":6,"sunday":6
}

def parse_layout_a(rows: List[List], headers: List[str], plan_start: date) -> List[Dict]:
    """Weeks as rows, days as columns."""
    sessions = []
    day_cols = {i: DAY_OFFSETS[str(h).lower().strip()]
                for i, h in enumerate(headers)
                if str(h).lower().strip() in DAY_OFFSETS}

    week_num = 0
    for row in rows:
        if not any(str(c).strip() for c in row if c is not None):
            continue
        
        first = str(row[0] or "").strip().lower()
        wk_match = re.search(r"\d+", first)
        if wk_match and any(kw in first for kw in ("week","wk","")):
            week_num = int(wk_match.group())
        else:
            week_num += 1

        week_start = plan_start + timedelta(weeks=week_num - 1)

        for col_idx, day_offset in day_cols.items():
            if col_idx >= len(row):
                continue
            cell = str(row[col_idx] or "").strip()
            if not cell or cell.lower() in ("rest","off","-",""):
                continue
            
            session_date = week_start + timedelta(days=day_offset)
            s = _parse_cell_to_session(cell, session_date)
            if s:
                sessions.append(s)
                
    return sessions


def parse_layout_b(rows: List[List], headers: List[str], plan_start: date) -> List[Dict]:
    """
    Weeks as columns, sessions as rows.

    Expected shape:
        Sport/Session  | W1       | W2       | W3       | W4
        Swim Mon       | 2x1500m  | 3x1000m  | 4x800m   | Easy 1km
        Bike Tue       | 2hr Z2   | 2.5hr Z2 | 3hr Z2   | 1hr recovery
        Run Wed        | 10km Z2  | 12km Z2  | 14km Z2  | 8km easy

    The first column labels the sport + day.  Subsequent columns are week cells.
    Week column headers must contain "w", "wk", "week", or a bare integer (1, 2 …).
    """
    if not headers or not rows:
        return []

    # ── Identify week columns ───────────────────────────────────────────────
    _WEEK_RE = re.compile(r"^(?:w(?:ee)?k?\.?\s*)?(\d+)$", re.IGNORECASE)

    week_cols: List[tuple] = []   # (col_index, week_number 0-based)
    for col_idx, hdr in enumerate(headers):
        hdr_str = str(hdr or "").strip()
        m = _WEEK_RE.match(hdr_str)
        if m:
            week_num = int(m.group(1)) - 1   # convert to 0-based
            week_cols.append((col_idx, week_num))

    if not week_cols:
        logger.warning("parse_layout_b: no week columns found in headers %s", headers)
        return []

    # ── Build sessions ──────────────────────────────────────────────────────
    sessions: List[Dict] = []

    for row in rows:
        if not any(str(c or "").strip() for c in row):
            continue   # blank row

        # First cell: "Swim Mon", "Bike Tuesday", "Run Wed", "Strength Fri" …
        first_cell = str(row[0] or "").strip()
        if not first_cell:
            continue

        # Extract day-of-week offset
        words = first_cell.lower().split()
        day_offset: Optional[int] = None
        for word in words:
            if word in DAY_OFFSETS:
                day_offset = DAY_OFFSETS[word]
                break

        # Extract sport hint from first cell
        sport_hint = _infer_sport(first_cell, "")

        for col_idx, week_num in week_cols:
            if col_idx >= len(row):
                continue
            cell = str(row[col_idx] or "").strip()
            if not cell or cell.lower() in ("rest", "off", "-", ""):
                continue

            week_start = plan_start + timedelta(weeks=week_num)

            # Use the detected day offset; fall back to Mon (0) when absent
            offset = day_offset if day_offset is not None else 0
            session_date = week_start + timedelta(days=offset)

            parsed = _parse_cell_to_session(cell, session_date)
            if parsed is None:
                continue

            # Override sport if the row label was more specific than the cell text
            if sport_hint != "cross_training":
                parsed["sport"] = sport_hint

            sessions.append(parsed)

    logger.info("parse_layout_b: extracted %d sessions", len(sessions))
    return sessions


def parse_layout_c(rows: List[List], headers: List[str], column_map: Optional[Dict] = None) -> List[Dict]:
    """Flat list with explicit date column."""
    ALIASES = {
        "date":     ["date","day"],
        "sport":    ["sport","activity","type","discipline"],
        "title":    ["title","name","workout","session"],
        "duration": ["duration","time","length"],
        "description": ["description","notes","details","desc","coach notes"],
        "tss":      ["tss","training stress"],
        "planned_if": ["if","intensity factor"],
    }
    col_idx = {}
    for i, h in enumerate(headers):
        key = str(h or "").lower().strip()
        for canonical, aliases in ALIASES.items():
            if key in aliases:
                col_idx[canonical] = i

    sessions = []
    for row in rows:
        if not any(str(c or "").strip() for c in row):
            continue
            
        def get_val(field):
            i = col_idx.get(field)
            return str(row[i] or "").strip() if i is not None and i < len(row) else ""

        raw_date = get_val("date")
        if not raw_date:
            continue
            
        try:
            from dateutil import parser as dp
            session_date = dp.parse(raw_date).date()
        except Exception:
            continue

        desc = get_val("description")
        sport_raw = get_val("sport")
        sport = _infer_sport(sport_raw, desc)
        
        sessions.append({
            "planned_date": session_date.isoformat(),
            "sport": sport,
            "title": get_val("title") or f"{sport.title()} session",
            "coaching_text": desc,
            "planned_duration_min": _parse_duration_str(get_val("duration")),
            "planned_tss": _safe_float(get_val("tss")),
            "planned_if": _safe_float(get_val("planned_if")),
        })
    return sessions


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------
def _parse_cell_to_session(cell_text: str, session_date: date) -> Optional[Dict]:
    """Parse a single grid cell into a session."""
    if not cell_text: return None
    lines = cell_text.splitlines()
    title = lines[0].strip()
    desc = "\n".join(lines[1:]).strip() if len(lines) > 1 else ""
    return {
        "planned_date": session_date.isoformat(),
        "title": title,
        "coaching_text": desc,
        "sport": _infer_sport(title, desc),
        "planned_duration_min": _parse_duration_str(title + " " + desc)
    }

def _infer_sport(sport_or_title: str, desc: str) -> str:
    combined = (sport_or_title + " " + desc).lower()
    if any(x in combined for x in ("swim", "pool")): return "swim"
    if any(x in combined for x in ("bike", "ride", "cycle", "trainerroad", "zwift")): return "bike"
    if any(x in combined for x in ("run", "jog", "treadmill")): return "run"
    if any(x in combined for x in ("strength", "gym", "lift")): return "strength"
    return "cross_training"

def _parse_duration_str(s: str) -> Optional[float]:
    if not s: return None
    s = s.lower()
    
    # e.g., "90min", "1.5hr", "45 m"
    m_min = re.search(r'(\d+)\s*(?:m|min|mins|minutes)', s)
    if m_min:
        return float(m_min.group(1))
        
    m_hr = re.search(r'(\d+(?:\.\d+)?)\s*(?:h|hr|hrs|hours)', s)
    if m_hr:
        return float(m_hr.group(1)) * 60
        
    return None

def _safe_float(val: str) -> Optional[float]:
    try:
        return float(val) if val else None
    except ValueError:
        return None

def _normalise_spreadsheet_session(s: Dict, athlete_id: str, file_name: str) -> Dict:
    import uuid
    uid = str(uuid.uuid4())
    s["session_id"] = f"spreadsheet_{uid[:8]}"
    s["source_platform"] = "spreadsheet"
    s["import_method"] = file_name
    
    # Default absent values
    s.setdefault("planned_tss", None)
    s.setdefault("planned_if", None)
    s.setdefault("planned_distance_m", None)
    s.setdefault("planned_elevation_m", None)
    s.setdefault("structure", {})
    s.setdefault("targets", {})
    return s


def _store_planned_session(session: Dict) -> None:
    sql = """
    INSERT INTO planned_sessions (
        session_id, source_platform, import_method, planned_date, sport, title,
        coaching_text, planned_duration_min, planned_tss, planned_if,
        planned_distance_m, planned_elevation_m, structure, targets
    ) VALUES (
        %(session_id)s, %(source_platform)s, %(import_method)s, %(planned_date)s, %(sport)s, %(title)s,
        %(coaching_text)s, %(planned_duration_min)s, %(planned_tss)s, %(planned_if)s,
        %(planned_distance_m)s, %(planned_elevation_m)s,
        %(structure)s::jsonb, %(targets)s::jsonb
    )
    ON CONFLICT (session_id) DO UPDATE SET
        title             = EXCLUDED.title,
        coaching_text     = EXCLUDED.coaching_text,
        planned_tss       = EXCLUDED.planned_tss,
        planned_duration_min = EXCLUDED.planned_duration_min,
        structure         = EXCLUDED.structure
    """
    import json
    # Ensure they are serialized strings if passing raw
    session["structure"] = json.dumps(session.get("structure", {}))
    session["targets"] = json.dumps(session.get("targets", {}))
    
    db.execute(sql, session)
